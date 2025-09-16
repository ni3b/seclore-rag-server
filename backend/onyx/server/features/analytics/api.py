from fastapi import APIRouter, Query, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import datetime

from typing import Tuple, List, Dict, Any

from onyx.auth.users import current_admin_user
from onyx.db.engine import get_session
from onyx.db.utils import parse_date_range
from onyx.db.models import User
from onyx.db.analytics import (
    fetch_bucketed_core_metrics,
    fetch_core_metrics_totals,
    fetch_assistant_usage_totals,
    fetch_kb_assistant_usage,
    fetch_user_assistant_usage,
)

router = APIRouter(prefix="/analytics", tags=["analytics"])

@router.get("")
def get_analytics(
    start: str = Query(...),
    end: str = Query(...),
    db_session: Session = Depends(get_session),
    user: User = Depends(current_admin_user),
):
    start_dt, end_dt = parse_date_range(start, end)

    totals = fetch_core_metrics_totals(db_session, start_dt, end_dt)
    bucketed = fetch_bucketed_core_metrics(db_session, start_dt, end_dt, bucket="day")
    assistant_usage_data = fetch_assistant_usage_totals(db_session, start_dt, end_dt)

    return {
        "totals": {
            "active_users": totals[0],
            "queries": totals[1],
            "input_tokens": totals[2],
            "output_tokens": totals[3],
            "likes": totals[4],
            "dislikes": totals[5],
        },
        "bucketed": bucketed,
        "assistant_data": assistant_usage_data,
    }


@router.get("/report", response_class=StreamingResponse)
def generate_analytics_report(
    start: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end: str = Query(..., description="End date in YYYY-MM-DD format"),
    db_session: Session = Depends(get_session),
    user: User = Depends(current_admin_user),
):
    """
    Generates and streams an XLSX report on assistant usage.

    The report contains two sheets:
    1.  **User Assistant Usage**: Pivoted metrics per user, with dynamic columns for each assistant.
    2.  **KB Assistant Usage**: Aggregated metrics per knowledge base and assistant.
    """
    start_dt, end_dt = parse_date_range(start, end)

    # Fetch datasets for the report using the new pivoted function
    user_data = fetch_user_assistant_usage(db_session, start_dt, end_dt)
    kb_data = fetch_kb_assistant_usage(db_session, start_dt, end_dt)

    # Create Excel workbook
    workbook = Workbook()

    # --- Sheet 1: User Assistant Usage ---
    ws1 = workbook.active
    ws1.title = "User Assistant Usage"

    if not user_data:
        ws1.append(["No user data available for the selected period."])
    else:
        headers = list(user_data[0].keys())
        ws1.append(headers)

        for row in user_data:
            ws1.append([row.get(header, "") for header in headers])

    # --- Sheet 2: KB Assistant Usage ---
    ws2 = workbook.create_sheet(title="KB Assistant Usage")

    if not kb_data:
        ws2.append(["No KB data available for the selected period."])
    else:
        headers = list(kb_data[0].keys())
        ws2.append(headers)

        for row in kb_data:
            ws2.append([row.get(header, "") for header in headers])

    # --- Formatting ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    def format_sheet(sheet, data):
        if not data:
            return
        # Format header row
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    format_sheet(ws1, user_data)
    format_sheet(ws2, kb_data)

    # --- Save to BytesIO stream ---
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)

    filename = f"astra_assistant_report_{start}_to_{end}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )